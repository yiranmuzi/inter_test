# -*- coding:utf-8 -*-
# @Time    : 2020/7/3 18:32
# @Author  : muzijie
# @File    : rw_excel.py

from openpyxl import load_workbook
def read_data(file_name,sheet_name):
    wb=load_workbook(file_name)
    data=wb[sheet_name]
    all_case=[]
    for i in range(2,data.max_row+1):
        case=[]
        for j in range(1,data.max_column-1):
            case.append(data.cell(i,j).value)
        all_case.append(case)
    return all_case
def write_data(file_name,sheet_name,row,colum,value):
    wb = load_workbook(file_name)
    data = wb[sheet_name]
    data.cell(row,colum).value=value
    wb.save(file_name)
if __name__ == '__main__':
    xx=read_data('test_case.xlsx','recharge')
    print(xx)
